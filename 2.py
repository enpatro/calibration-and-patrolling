from pathlib import Path
import csv, json, zipfile, shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

base = Path('/mnt/data/Compliance_Portal_Firebase_GitHub_Bundle_v5')
if base.exists(): shutil.rmtree(base)
for d in ['assets/css','assets/js','data','templates','tools','certificates','firebase']:
    (base/d).mkdir(parents=True, exist_ok=True)

months = ['Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec','Jan','Feb','Mar']
weeks = [f'{m} W{w}' for m in months for w in range(1,5)]
cal_headers = ['area','machineNo','gaugeId','certificateNo','gaugeDescription','range','workRange','lc','errorPercent','calibrationDate','dueDate','error %','certificateFileName']
cal_rows = [
 ['CA4','CA4M28','CA4M28-PG02','VI/25-26/4650-122','SHUTTER DOOR','0-0.4 MPA','0.1-0.2 MPA','0.11','0.05 MPA','27-12-2025','28-12-2026','0.002','CA4M28-PG02.pdf'],
 ['CA4','CA4M28','CA4M28-PG03','VI/25-26/4650-123','FRL UNIT','0-1 MPA','0.4-0.6 MPA','0.11','0.05 MPA','27-12-2025','28-12-2026','0.002','CA4M28-PG03.pdf'],
 ['CA4','CA4M28','CA4M28-PG04','VI/25-26/4650-124','LUBRICATION','0-5 MPA','2-3 MPA','0.11','0.2 MPA','27-12-2025','28-12-2026','0.002','CA4M28-PG04.pdf'],
 ['CA4','CA4M28','CA4M28-PG05','VI/25-26/4650-125','HYDRAULIC UNIT','0-10 MPA','2-5 MPA','0.11','0.5 MPA','27-12-2025','28-12-2026','0.002','CA4M28-PG05.pdf'],
 ['CA4','CA4M28','CA4M28-PG06','VI/25-26/4650-126','TOOL PAT ADV / RTN','0-6 MPA','1-3 MPA','0.11','0.1 MPA','27-12-2025','28-12-2026','0.002','CA4M28-PG06.pdf'],
 ['CA4','CA4M28','CA4M28-PG07','VI/25-26/4650-127','WORK CLAMP UNCLAMP','0-6 MPA','2-4 MPA','0.11','0.1 MPA','27-12-2025','28-12-2026','0.002','CA4M28-PG07.pdf']
]
pat_headers = ['Machine','Sl.No','Check Points','Checking Method'] + weeks
pat_rows = [
 ['CA04M01','1','Operator panel push buttons, indication lamps, selector s/w mounting should be proper tight.','By hand'] + ['X','O','O','O','O','O','O','O','X','O','X',''] + ['']*(len(weeks)-12),
 ['CA04M01','2','There should be no damage of lamps, switches and push buttons.','Visual'] + ['O','O','X','O','O','X','O','O','','','',''] + ['']*(len(weeks)-12),
 ['CA04M01','3','Front doors, side doors and safety lock switch mounting condition should be proper.','By hand'] + ['O','O','O','O','O','O','O','O','','','',''] + ['']*(len(weeks)-12),
 ['CA04M02','1','Machine lamp and tower lamp condition check.','Visual'] + ['']*len(weeks),
]
cal_data = [dict(zip(cal_headers,r)) for r in cal_rows]
pat_data = [dict(zip(pat_headers,r)) for r in pat_rows]
(base/'data/calibration.json').write_text(json.dumps(cal_data, indent=2, ensure_ascii=False), encoding='utf-8')
(base/'data/patrolling.json').write_text(json.dumps(pat_data, indent=2, ensure_ascii=False), encoding='utf-8')

with open(base/'templates/calibration_template.csv','w',newline='',encoding='utf-8-sig') as f:
    w=csv.writer(f); w.writerow(cal_headers); w.writerows(cal_rows)
with open(base/'templates/patrolling_template.csv','w',newline='',encoding='utf-8-sig') as f:
    w=csv.writer(f); w.writerow(pat_headers); w.writerows(pat_rows)

thin = Side(style='thin', color='CBD5E1')
def style(ws, pat=False):
    ws.freeze_panes='A2'
    for row in ws.iter_rows():
        for c in row:
            c.border=Border(bottom=thin, right=thin); c.alignment=Alignment(vertical='center', wrap_text=True)
    for col in ws.columns:
        ml = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(ml+2,10),55)

wb=Workbook(); ws=wb.active; ws.title='Calibration'; ws.append(cal_headers)
for r in cal_rows: ws.append(r)
for c in ws[1]: c.fill=PatternFill('solid', fgColor='1F4E79'); c.font=Font(color='FFFFFF', bold=True); c.alignment=Alignment(horizontal='center')
style(ws); wb.save(base/'templates/calibration_template.xlsx')
wb=Workbook(); ws=wb.active; ws.title='Patrolling'; ws.append(pat_headers)
for r in pat_rows: ws.append(r)
for idx,c in enumerate(ws[1],1):
    h=str(c.value); color='4472C4' if idx<=4 else '5B9BD5' if h.startswith('Apr') else 'F4B183' if h.startswith('May') else '63C66D' if h.startswith('Jun') else '70AD47'
    c.fill=PatternFill('solid', fgColor=color); c.font=Font(color='FFFFFF', bold=True); c.alignment=Alignment(horizontal='center')
dv=DataValidation(type='list', formula1='"O,X"', allow_blank=True); ws.add_data_validation(dv); dv.add('E2:AZ10000')
style(ws); wb.save(base/'templates/patrolling_template.xlsx')

css = r''':root{--primary:#2563eb;--success:#16a34a;--warning:#f59e0b;--danger:#dc2626;--text:#172033;--muted:#64748b;--border:#e2e8f0;--shadow:0 14px 35px rgba(15,23,42,.09);--radius:18px}*{box-sizing:border-box}body{margin:0;font-family:Inter,Segoe UI,Arial,sans-serif;background:linear-gradient(135deg,#eef4ff,#f8fafc 45%,#eefdf6);color:var(--text)}header{position:sticky;top:0;background:rgba(255,255,255,.94);backdrop-filter:blur(12px);border-bottom:1px solid var(--border);z-index:10}.topbar{max-width:1580px;margin:auto;padding:14px 20px;display:flex;justify-content:space-between;gap:12px;align-items:center}.brand{display:flex;gap:12px;align-items:center}.logo{width:46px;height:46px;background:linear-gradient(135deg,var(--primary),#22c55e);border-radius:15px;color:white;display:grid;place-items:center;font-weight:950}.brand h1{font-size:20px;margin:0}.brand p{margin:3px 0 0;color:var(--muted);font-size:12px}.tabs{display:flex;gap:8px;flex-wrap:wrap}.tab,button{border:1px solid var(--border);background:#fff;border-radius:999px;padding:10px 14px;font-weight:800;cursor:pointer;text-decoration:none;color:var(--text)}.tab.active,.tab:hover,button.primary{background:var(--primary);border-color:var(--primary);color:#fff}button.danger{background:var(--danger);color:#fff;border-color:var(--danger)}main{max-width:1580px;margin:auto;padding:20px}.section{display:none}.section.active{display:block}.panel,.card{background:rgba(255,255,255,.96);border:1px solid var(--border);border-radius:var(--radius);box-shadow:var(--shadow)}.panel{padding:20px;margin-bottom:18px}.title{font-size:clamp(24px,3vw,38px);margin:0 0 8px;letter-spacing:-.04em}.sub{color:var(--muted);margin:0;line-height:1.5}.kpi-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:18px}.kpi{padding:18px;position:relative;overflow:hidden}.icon{width:42px;height:42px;border-radius:13px;background:#eff6ff;display:grid;place-items:center;font-size:21px;margin-bottom:10px}.kpi h3{font-size:12px;color:var(--muted);margin:0;text-transform:uppercase}.value{font-size:30px;font-weight:950;margin-top:6px}.note{font-size:12px;color:var(--muted);margin-top:4px}.toolbar{display:grid;grid-template-columns:repeat(5,1fr);gap:12px;margin-top:15px}input,select{width:100%;padding:12px;border:1px solid var(--border);border-radius:14px;outline:none;background:#fff}.grid-2{display:grid;grid-template-columns:1fr 1fr;gap:18px}.grid-3{display:grid;grid-template-columns:repeat(3,1fr);gap:18px}.full-row{grid-column:1/-1}.upload-grid{display:grid;grid-template-columns:1fr 1fr;gap:18px}.card-header{padding:16px 18px 0;display:flex;gap:12px;justify-content:space-between;align-items:center}.card-header h2{font-size:17px;margin:0}.card-body{padding:16px 18px 18px}.table-wrap{overflow:auto;max-height:560px;border:1px solid var(--border);border-radius:16px;background:#fff}table{width:100%;border-collapse:collapse;min-width:1030px;font-size:13px}th,td{padding:10px 11px;border-bottom:1px solid var(--border);white-space:nowrap;text-align:left}th{position:sticky;top:0;background:#f8fafc;z-index:2;color:var(--muted);font-size:12px;text-transform:uppercase}.badge{display:inline-flex;border-radius:999px;padding:5px 9px;font-size:12px;font-weight:900}.ok{background:#dcfce7;color:#166534}.warn{background:#fef3c7;color:#92400e}.danger-b{background:#fee2e2;color:#991b1b}.neutral{background:#e2e8f0;color:#334155}.week-ok{color:var(--success);font-weight:950;text-align:center}.week-x{color:var(--danger);background:#fee2e2;font-weight:950;text-align:center}.week-blank{color:#94a3b8;text-align:center}.dropzone{border:2px dashed #93c5fd;border-radius:18px;background:#eff6ff;padding:18px}.actions{display:flex;gap:10px;flex-wrap:wrap;margin-top:14px}.mini{font-size:12px;color:var(--muted)}.pager{display:flex;justify-content:flex-end;gap:8px;align-items:center;margin-top:12px}canvas{width:100%;height:260px;display:block}.year-strip{display:grid;grid-template-columns:repeat(12,1fr);gap:8px;margin:12px 0}.month-box{min-height:72px;border-radius:14px;padding:10px;text-align:center;border:1px solid var(--border);background:#f8fafc}.month-box .m{font-size:12px;font-weight:900;color:#475569}.month-box .p{font-size:22px;font-weight:950;margin-top:5px}.month-box.good{background:#dcfce7;color:#166534}.month-box.medium{background:#fef3c7;color:#92400e}.month-box.bad{background:#fee2e2;color:#991b1b}.ng-dashboard{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:16px}.ng-tile{border-radius:18px;padding:16px;border:1px solid var(--border);background:linear-gradient(135deg,#fff,#f8fafc);box-shadow:0 10px 24px rgba(15,23,42,.07)}.ng-tile.danger{background:linear-gradient(135deg,#fff,#fff1f2);border-color:#fecaca}.ng-tile.oktile{background:linear-gradient(135deg,#fff,#f0fdf4);border-color:#bbf7d0}.ng-tile.warnTile{background:linear-gradient(135deg,#fff,#fffbeb);border-color:#fde68a}.ng-icon{font-size:28px;margin-bottom:8px}.ng-title{font-size:12px;color:var(--muted);font-weight:900;text-transform:uppercase}.ng-value{font-size:28px;font-weight:950}.ng-caption{font-size:12px;color:var(--muted)}.issue-card-wrap{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:14px}.issue-card{border:1px solid var(--border);border-left:6px solid var(--danger);border-radius:16px;padding:13px;background:#fff}.issue-card h3{font-size:14px;margin:0 0 6px}.issue-card p{font-size:12px;color:var(--muted);margin:0}.issue-meta{display:flex;gap:8px;flex-wrap:wrap;margin-top:9px}.issue-meta span{border-radius:999px;background:#fee2e2;color:#991b1b;font-size:11px;font-weight:900;padding:4px 8px}.authbox{display:flex;gap:8px;align-items:center;flex-wrap:wrap}.authbox input{width:190px}.firebase-status{font-size:12px;color:var(--muted);font-weight:800}@media(max-width:1050px){.kpi-grid{grid-template-columns:repeat(2,1fr)}.grid-2,.grid-3,.upload-grid{grid-template-columns:1fr}.toolbar{grid-template-columns:repeat(2,1fr)}.ng-dashboard{grid-template-columns:repeat(2,1fr)}.issue-card-wrap{grid-template-columns:1fr}.year-strip{grid-template-columns:repeat(4,1fr)}}@media(max-width:650px){.topbar{flex-direction:column;align-items:flex-start}.kpi-grid,.toolbar,.ng-dashboard{grid-template-columns:1fr}.year-strip{grid-template-columns:repeat(2,1fr)}main{padding:14px}.authbox input{width:100%}}
'''
(base/'assets/css/style.css').write_text(css, encoding='utf-8')

index = '''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Compliance Portal | Firebase + GitHub Pages</title>
<link rel="stylesheet" href="assets/css/style.css">
</head>
<body>
<header><div class="topbar"><div class="brand"><div class="logo">CP</div><div><h1>Compliance Portal</h1><p>Calibration + Patrolling | GitHub Pages + Firebase Firestore + Storage</p></div></div><nav class="tabs"><button class="tab active" data-tab="dashboard">Dashboard</button><button class="tab" data-tab="calibration">Calibration</button><button class="tab" data-tab="patrolling">Patrolling</button><button class="tab" data-tab="bulk">Bulk/Admin</button><button class="tab" data-tab="setup">Firebase Setup</button></nav></div></header>
<main>
<section id="dashboard" class="section active"><div class="panel"><h1 class="title">Unified compliance control tower</h1><p class="sub">Firebase-ready portal. If Firebase config is filled and login is done, live Firestore data loads. If not configured, portal uses local JSON sample data from GitHub.</p><div class="actions"><span id="dataMode" class="badge neutral">Checking data mode...</span><span id="userBadge" class="badge neutral">Not logged in</span></div></div><div class="kpi-grid" id="dashboardKpis"></div><div class="grid-2"><div class="card"><div class="card-header"><h2>Monthly Calibration Compliance</h2><span class="badge ok" id="calComplianceBadge">0%</span></div><div class="card-body"><canvas id="calMonthlyChart"></canvas></div></div><div class="card"><div class="card-header"><h2>Full Year Patrolling Compliance - Apr to Mar</h2><span class="badge ok" id="patComplianceBadge">0%</span></div><div class="card-body"><div class="year-strip" id="patYearStrip"></div><canvas id="patMonthlyChart"></canvas></div></div></div><div class="grid-3" style="margin-top:18px"><div class="card full-row"><div class="card-header"><h2>Open Abnormal Report - Latest Check is X</h2><span class="badge danger-b" id="openAbnCount">0 open</span></div><div class="card-body"><div class="table-wrap"><table><thead><tr><th>Machine</th><th>Sl.No</th><th>Reported Issue / Check Point</th><th>Method</th><th>Last Check Week</th><th>Last Result</th><th>Previous O Week</th></tr></thead><tbody id="openAbnormalTable"></tbody></table></div></div></div><div class="card full-row"><div class="card-header"><h2>Abnormal to OKIE Conversion Timeline</h2><span class="badge ok" id="conversionCount">0 closed</span></div><div class="card-body"><div class="table-wrap"><table><thead><tr><th>Machine</th><th>Sl.No</th><th>Reported Issue</th><th>Abnormal Week</th><th>OK Week</th><th>Closure Time</th><th>Latest Status</th></tr></thead><tbody id="conversionTable"></tbody></table></div></div></div><div class="card full-row"><div class="card-header"><h2>Not Done Machine Details</h2><span class="badge neutral" id="notDoneCount">0 not done</span></div><div class="card-body"><div class="table-wrap"><table><thead><tr><th>Machine</th><th>Sl.No</th><th>Check Point</th><th>Method</th><th>Status</th><th>Action Required</th></tr></thead><tbody id="notDoneTable"></tbody></table></div></div></div><div class="card full-row"><div class="card-header"><h2>Top NG Points - Reported Issue Wise</h2><span class="badge danger-b">Latest X only</span></div><div class="card-body"><div class="ng-dashboard" id="ngSummaryTiles"></div><div class="issue-card-wrap" id="topIssueCards"></div><canvas id="topIssuesChart"></canvas></div></div><div class="card"><div class="card-header"><h2>Top Overdue Gauges</h2></div><div class="card-body"><div class="table-wrap"><table><tbody id="topOverdueTable"></tbody></table></div></div></div><div class="card"><div class="card-header"><h2>Area-wise Compliance</h2></div><div class="card-body"><canvas id="areaChart"></canvas></div></div></div></section>
<section id="calibration" class="section"><div class="panel"><h1 class="title">Calibration Management</h1><div class="toolbar"><input id="gaugeSearch" placeholder="Search Gauge ID / Description"><input id="machineSearchCal" placeholder="Search Machine No"><input id="areaSearch" placeholder="Search Area"><input id="certificateSearch" placeholder="Search Certificate No"><select id="statusFilter"><option value="">All status</option><option>Valid</option><option>Due Soon</option><option>Overdue</option></select></div></div><div class="card"><div class="card-header"><h2>Calibration Records</h2><span class="badge neutral" id="calResultCount">0 records</span></div><div class="card-body"><div class="table-wrap"><table><thead><tr><th>Area</th><th>Machine No</th><th>Gauge ID</th><th>Certificate No</th><th>Description</th><th>Range</th><th>Work Range</th><th>LC</th><th>Error Percent</th><th>Cal. Date</th><th>Due Date</th><th>Error %</th><th>Status</th><th>Certificate</th></tr></thead><tbody id="calTable"></tbody></table></div><div class="pager"><button id="calPrev">Prev</button><span id="calPageInfo"></span><button class="primary" id="calNext">Next</button></div></div></div></section>
<section id="patrolling" class="section"><div class="panel"><h1 class="title">Patrolling Management</h1><p class="sub">Open abnormal is shown only when latest non-blank check is X. If latest check is O, it is OKIE.</p><div class="toolbar"><input id="machineSearchPat" placeholder="Search Machine"><input id="checkpointSearch" placeholder="Search Check Point"><select id="patStatusFilter"><option value="">All latest status</option><option value="OPEN_X">Open abnormal only</option><option value="OK">Latest OK</option><option value="NOT_DONE">Not done</option></select><select id="monthFilter"><option value="">All months</option></select><button class="primary" id="toggleWeeks">Toggle All Weeks</button></div></div><div class="card"><div class="card-header"><h2>Patrolling Checkpoints</h2><span class="badge neutral" id="patResultCount">0 records</span></div><div class="card-body"><div class="table-wrap"><table><thead id="patHead"></thead><tbody id="patTable"></tbody></table></div><div class="pager"><button id="patPrev">Prev</button><span id="patPageInfo"></span><button class="primary" id="patNext">Next</button></div></div></div></section>
<section id="bulk" class="section"><div class="panel"><h1 class="title">Bulk/Admin</h1><p class="sub">Upload CSV/JSON to browser, export JSON, or sync data to Firebase Firestore after login. Certificate PDFs can be uploaded to Firebase Storage.</p><div class="authbox"><input id="email" placeholder="Admin email"><input id="password" type="password" placeholder="Password"><button class="primary" id="loginBtn">Login</button><button id="logoutBtn">Logout</button><span id="firebaseStatus" class="firebase-status">Firebase status...</span></div><div class="actions"><a class="tab" href="templates/calibration_template.xlsx">Calibration Excel Template</a><a class="tab" href="templates/patrolling_template.xlsx">Patrolling Excel Template</a><button class="primary" id="downloadCalCsv">Download Calibration CSV</button><button class="primary" id="downloadPatCsv">Download Patrolling CSV</button></div></div><div class="upload-grid"><div class="card"><div class="card-header"><h2>Calibration Bulk Upload</h2><span id="calUploadStatus" class="badge neutral">Ready</span></div><div class="card-body"><div class="dropzone"><input id="calUpload" type="file" accept=".csv,.json"><p class="mini">Upload calibration CSV/JSON. Then export JSON or sync to Firestore.</p></div><div class="actions"><button class="primary" id="exportCalJson">Export calibration.json</button><button class="primary" id="syncCalFirestore">Sync Calibration to Firestore</button></div></div></div><div class="card"><div class="card-header"><h2>Patrolling Bulk Upload</h2><span id="patUploadStatus" class="badge neutral">Ready</span></div><div class="card-body"><div class="dropzone"><input id="patUpload" type="file" accept=".csv,.json"><p class="mini">Upload patrolling CSV/JSON. Values allowed: O, X or blank.</p></div><div class="actions"><button class="primary" id="exportPatJson">Export patrolling.json</button><button class="primary" id="syncPatFirestore">Sync Patrolling to Firestore</button></div></div></div></div><div class="card" style="margin-top:18px"><div class="card-header"><h2>Certificate PDF Upload to Firebase Storage</h2></div><div class="card-body"><input id="certUpload" type="file" accept=".pdf" multiple><div class="actions"><button class="primary" id="uploadCerts">Upload Certificates</button></div><p class="mini">PDFs are saved under Firebase Storage folder: certificates/fileName.pdf</p></div></div><div class="card" style="margin-top:18px"><div class="card-header"><h2>Upload / Sync Log</h2></div><div class="card-body"><pre id="validationLog" style="white-space:pre-wrap;background:#f8fafc;border:1px solid var(--border);border-radius:14px;padding:14px;max-height:260px;overflow:auto">No action yet.</pre></div></div></section>
<section id="setup" class="section"><div class="panel"><h1 class="title">Firebase Setup Checklist</h1><p class="sub">Use this section after creating your Firebase project. Fill assets/js/firebase-config.js with your Firebase web app config.</p></div><div class="grid-2"><div class="card"><div class="card-header"><h2>Required Firebase Services</h2></div><div class="card-body"><ul><li>Authentication: enable Email/Password</li><li>Firestore Database: create in production mode</li><li>Storage: enable for certificate PDFs</li><li>Authorized domain: add your GitHub Pages domain</li></ul></div></div><div class="card"><div class="card-header"><h2>Firestore Collections</h2></div><div class="card-body"><pre>calibration/{gaugeId}
patrolling/{machine_slNo}
auditLogs/{autoId}
users/{uid}</pre></div></div></div></section>
</main>
<script type="module" src="assets/js/app.js"></script>
</body>
</html>
'''
(base/'index.html').write_text(index, encoding='utf-8')

config_js = '''// Firebase web config. Replace these values from Firebase Console > Project Settings > Your apps > Web app.
export const firebaseConfig = {
  apiKey: "REPLACE_WITH_API_KEY",
  authDomain: "REPLACE_WITH_PROJECT_ID.firebaseapp.com",
  projectId: "REPLACE_WITH_PROJECT_ID",
  storageBucket: "REPLACE_WITH_PROJECT_ID.appspot.com",
  messagingSenderId: "REPLACE_WITH_SENDER_ID",
  appId: "REPLACE_WITH_APP_ID"
};
'''
(base/'assets/js/firebase-config.js').write_text(config_js, encoding='utf-8')
(base/'assets/js/firebase-config.example.js').write_text(config_js, encoding='utf-8')

app_js = r'''import { firebaseConfig } from './firebase-config.js';
import { initializeApp } from 'https://www.gstatic.com/firebasejs/12.0.0/firebase-app.js';
import { getAuth, signInWithEmailAndPassword, signOut, onAuthStateChanged } from 'https://www.gstatic.com/firebasejs/12.0.0/firebase-auth.js';
import { getFirestore, collection, getDocs, doc, writeBatch, serverTimestamp, addDoc } from 'https://www.gstatic.com/firebasejs/12.0.0/firebase-firestore.js';
import { getStorage, ref, uploadBytes, getDownloadURL } from 'https://www.gstatic.com/firebasejs/12.0.0/firebase-storage.js';

const WEEKS=['Apr W1','Apr W2','Apr W3','Apr W4','May W1','May W2','May W3','May W4','Jun W1','Jun W2','Jun W3','Jun W4','Jul W1','Jul W2','Jul W3','Jul W4','Aug W1','Aug W2','Aug W3','Aug W4','Sep W1','Sep W2','Sep W3','Sep W4','Oct W1','Oct W2','Oct W3','Oct W4','Nov W1','Nov W2','Nov W3','Nov W4','Dec W1','Dec W2','Dec W3','Dec W4','Jan W1','Jan W2','Jan W3','Jan W4','Feb W1','Feb W2','Feb W3','Feb W4','Mar W1','Mar W2','Mar W3','Mar W4'];
const MONTHS=['Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec','Jan','Feb','Mar'];
const CAL_HEADERS=['area','machineNo','gaugeId','certificateNo','gaugeDescription','range','workRange','lc','errorPercent','calibrationDate','dueDate','error %','certificateFileName'];
const PAT_HEADERS=['Machine','Sl.No','Check Points','Checking Method',...WEEKS];
const state={calibration:[],patrolling:[],filteredCal:[],filteredPat:[],calPage:1,patPage:1,pageSize:50,user:null};
let showAllWeeks=false, fb=null;
const isFirebaseConfigured = firebaseConfig.apiKey && !firebaseConfig.apiKey.includes('REPLACE');
if(isFirebaseConfigured){const app=initializeApp(firebaseConfig);fb={auth:getAuth(app),db:getFirestore(app),storage:getStorage(app)};}

const $=id=>document.getElementById(id);
function esc(s){return String(s??'').replace(/[&<>"]/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[m]))}
function log(s){$('validationLog').textContent=new Date().toLocaleString()+"\n"+s+"\n\n"+$('validationLog').textContent}
function pct(n,d){return d?Math.round(n/d*100):0}
function uniq(a){return[...new Set(a.filter(Boolean))].sort()}
function parseDate(v){if(!v)return null;if(/^\d{2}-\d{2}-\d{4}$/.test(v)){let[d,m,y]=v.split('-');return new Date(`${y}-${m}-${d}`)}let d=new Date(v);return isNaN(d)?null:d}
function fmt(v){let d=parseDate(v);return d?d.toLocaleDateString('en-IN',{day:'2-digit',month:'short',year:'numeric'}):''}
function debounce(fn,ms){let t;return(...a)=>{clearTimeout(t);t=setTimeout(()=>fn(...a),ms)}}

async function loadJson(p){try{let r=await fetch(p,{cache:'no-store'});return r.ok?await r.json():[]}catch(e){return[]}}
async function loadFirestoreCollection(name){if(!fb||!state.user)return null;let snap=await getDocs(collection(fb.db,name));return snap.docs.map(d=>d.data())}
async function loadData(){let fromFirebase=false;try{if(fb&&state.user){let cal=await loadFirestoreCollection('calibration');let pat=await loadFirestoreCollection('patrolling');if(cal?.length||pat?.length){state.calibration=(cal||[]).map(normCal);state.patrolling=(pat||[]).map(normPat);fromFirebase=true}}}catch(e){log('Firestore load failed, using JSON fallback: '+e.message)}if(!fromFirebase){let [cal,pat]=await Promise.all([loadJson('data/calibration.json'),loadJson('data/patrolling.json')]);state.calibration=cal.map(normCal);state.patrolling=pat.map(normPat)}state.filteredCal=state.calibration;state.filteredPat=state.patrolling;$('dataMode').textContent=fromFirebase?'Live Firebase Firestore':'GitHub JSON fallback';$('firebaseStatus').textContent=isFirebaseConfigured?'Firebase configured':'Firebase config pending';renderAll()}
function normCal(r){let due=parseDate(r.dueDate);let daysLeft=due?Math.ceil((due-new Date())/86400000):null;let status=daysLeft!==null&&daysLeft<0?'Overdue':daysLeft!==null&&daysLeft<=30?'Due Soon':'Valid';return{...r,status,daysLeft}}
function normPat(r){let o={Machine:r.Machine||r.machine||'', 'Sl.No':r['Sl.No']||r.slNo||'', 'Check Points':r['Check Points']||r['Check Point']||r.checkPoint||'', 'Checking Method':r['Checking Method']||r.checkingMethod||''};WEEKS.forEach(w=>o[w]=(r[w]||'').toString().trim().toUpperCase());return o}
function latestStatus(r,wks=WEEKS){let lastWeek='',lastValue='';for(let i=wks.length-1;i>=0;i--){let v=r[wks[i]];if(v){lastWeek=wks[i];lastValue=v;break}}let lastX='',okAfterX='',prevO='';for(let i=0;i<wks.length;i++){let v=r[wks[i]];if(v==='O'&&!lastX)prevO=wks[i];if(v==='X'){lastX=wks[i];okAfterX=''}if(lastX&&v==='O')okAfterX=wks[i]}let status=!lastValue?'NOT_DONE':lastValue==='X'?'OPEN_X':'OK';return{status,lastWeek,lastValue,lastX,okAfterX,prevO}}
function reportData(){let open=[],converted=[],notdone=[];state.patrolling.forEach(r=>{let s=latestStatus(r);if(s.status==='OPEN_X')open.push({...r,_ls:s});if(s.status==='NOT_DONE')notdone.push({...r,_ls:s});if(s.okAfterX&&s.status==='OK')converted.push({...r,_ls:s})});return{open,converted,notdone}}
function closureWeeks(x,o){let xi=WEEKS.indexOf(x),oi=WEEKS.indexOf(o);return xi>=0&&oi>=0&&oi>=xi?`${oi-xi} week(s)`:''}
function renderAll(){renderKpis();renderCalTable();renderPatHeader();renderPatTable();renderReports();renderCharts();renderTopOverdue();renderYearStrip()}
function renderKpis(){let rep=reportData(),tg=state.calibration.length,due=state.calibration.filter(x=>x.status==='Due Soon').length,ov=state.calibration.filter(x=>x.status==='Overdue').length,tm=uniq(state.patrolling.map(x=>x.Machine)).length,tc=state.patrolling.length,cc=pct(tg-ov,tg),pc=patLatestCompliance(),overall=Math.round((cc+pc)/2);let arr=[['📏','Total Gauges',tg,'Calibration records'],['⏳','Due Soon',due,'Within 30 days'],['🚨','Overdue',ov,'Need action'],['🏭','Machines',tm,'Patrolling scope'],['✅','Checkpoints',tc,'Inspection items'],['⚠️','Open NG',rep.open.length,'Latest check is X'],['📌','Not Done',rep.notdone.length,'No check entry'],['🛡️','Overall',overall+'%','Combined score']];$('dashboardKpis').innerHTML=arr.map(k=>`<div class="card kpi"><div class="icon">${k[0]}</div><h3>${k[1]}</h3><div class="value">${k[2]}</div><div class="note">${k[3]}</div></div>`).join('');$('calComplianceBadge').textContent=cc+'%';$('patComplianceBadge').textContent=pc+'%'}
function renderCalTable(){let s=(state.calPage-1)*state.pageSize,rows=state.filteredCal.slice(s,s+state.pageSize);$('calResultCount').textContent=state.filteredCal.length+' records';$('calTable').innerHTML=rows.map(r=>{let b=r.status==='Overdue'?'danger-b':r.status==='Due Soon'?'warn':'ok';let cert=r.certificateUrl||(r.certificateFileName?`certificates/${r.certificateFileName}`:'');return`<tr><td>${esc(r.area)}</td><td><b>${esc(r.machineNo)}</b></td><td><b>${esc(r.gaugeId)}</b></td><td>${esc(r.certificateNo)}</td><td>${esc(r.gaugeDescription)}</td><td>${esc(r.range)}</td><td>${esc(r.workRange)}</td><td>${esc(r.lc)}</td><td>${esc(r.errorPercent)}</td><td>${fmt(r.calibrationDate)}</td><td>${fmt(r.dueDate)}</td><td>${esc(r['error %']||r.errorPct||'')}</td><td><span class="badge ${b}">${r.status}</span></td><td>${cert?`<a href="${cert}" target="_blank">View PDF</a>`:'NA'}</td></tr>`}).join('');pager('cal',state.filteredCal.length,state.calPage)}
function renderPatHeader(){let cw=showAllWeeks?WEEKS:WEEKS.slice(0,12);$('patHead').innerHTML=`<tr><th>Machine</th><th>Sl.No</th><th>Check Points</th><th>Checking Method</th>${cw.map(w=>`<th>${w}</th>`).join('')}<th>Latest Status</th><th>Last Week</th></tr>`}
function renderPatTable(){let s=(state.patPage-1)*state.pageSize,rows=state.filteredPat.slice(s,s+state.pageSize),cw=showAllWeeks?WEEKS:WEEKS.slice(0,12);$('patResultCount').textContent=state.filteredPat.length+' records';$('patTable').innerHTML=rows.map(r=>{let ls=latestStatus(r),b=ls.status==='OPEN_X'?'danger-b':ls.status==='OK'?'ok':'neutral',label=ls.status==='OPEN_X'?'ABNORMAL':ls.status==='OK'?'OKIE':'NOT DONE';return`<tr><td><b>${esc(r.Machine)}</b></td><td>${esc(r['Sl.No'])}</td><td>${esc(r['Check Points'])}</td><td>${esc(r['Checking Method'])}</td>${cw.map(w=>weekCell(r[w])).join('')}<td><span class="badge ${b}">${label}</span></td><td>${esc(ls.lastWeek||'-')}</td></tr>`}).join('');pager('pat',state.filteredPat.length,state.patPage)}
function renderReports(){let rep=reportData();$('openAbnCount').textContent=rep.open.length+' open';$('conversionCount').textContent=rep.converted.length+' closed';$('notDoneCount').textContent=rep.notdone.length+' not done';$('openAbnormalTable').innerHTML=rep.open.length?rep.open.map(r=>`<tr><td><b>${esc(r.Machine)}</b></td><td>${esc(r['Sl.No'])}</td><td>${esc(r['Check Points'])}</td><td>${esc(r['Checking Method'])}</td><td>${esc(r._ls.lastWeek)}</td><td><span class="badge danger-b">X</span></td><td>${esc(r._ls.prevO||'-')}</td></tr>`).join(''):'<tr><td colspan="7">No open abnormal.</td></tr>';$('conversionTable').innerHTML=rep.converted.length?rep.converted.map(r=>`<tr><td><b>${esc(r.Machine)}</b></td><td>${esc(r['Sl.No'])}</td><td>${esc(r['Check Points'])}</td><td><span class="badge danger-b">${esc(r._ls.lastX)}</span></td><td><span class="badge ok">${esc(r._ls.okAfterX)}</span></td><td>${closureWeeks(r._ls.lastX,r._ls.okAfterX)}</td><td><span class="badge ok">OKIE</span></td></tr>`).join(''):'<tr><td colspan="7">No abnormal to OKIE conversion.</td></tr>';$('notDoneTable').innerHTML=rep.notdone.length?rep.notdone.map(r=>`<tr><td><b>${esc(r.Machine)}</b></td><td>${esc(r['Sl.No'])}</td><td>${esc(r['Check Points'])}</td><td>${esc(r['Checking Method'])}</td><td><span class="badge neutral">NOT DONE</span></td><td>Complete patrolling entry and update O / X</td></tr>`).join(''):'<tr><td colspan="6">No not-done checkpoints.</td></tr>';renderNgVisuals()}
function renderNgVisuals(){let rep=reportData(),ok=state.patrolling.filter(r=>latestStatus(r).status==='OK').length;let tiles=[['🚨','Open NG Points',rep.open.length,'Latest check is X','danger'],['✅','Converted to OKIE',rep.converted.length,'Earlier X closed by O','oktile'],['📌','Not Done Points',rep.notdone.length,'No weekly check entry','warnTile'],['🛡️','Latest OK Points',ok,'Current healthy points','oktile']];$('ngSummaryTiles').innerHTML=tiles.map(x=>`<div class="ng-tile ${x[4]}"><div class="ng-icon">${x[0]}</div><div class="ng-title">${x[1]}</div><div class="ng-value">${x[2]}</div><div class="ng-caption">${x[3]}</div></div>`).join('');let items=rep.open.slice(0,6);$('topIssueCards').innerHTML=items.length?items.map(r=>`<div class="issue-card"><h3>${esc(r['Check Points'])}</h3><p>${esc(r['Checking Method'])}</p><div class="issue-meta"><span>${esc(r.Machine)}</span><span>Sl.No ${esc(r['Sl.No'])}</span><span>${esc(r._ls.lastWeek)}</span></div></div>`).join(''):'<div class="issue-card"><h3>No open NG points</h3><p>All latest patrolling checks are OKIE or not done.</p></div>'}
function renderYearStrip(){let data=monthlyPat();$('patYearStrip').innerHTML=data.map(d=>{let cls=d.value>=95?'good':d.value>=80?'medium':'bad';return`<div class="month-box ${cls}"><div class="m">${d.label}</div><div class="p">${d.value}%</div></div>`}).join('')}
function renderCharts(){bar('calMonthlyChart',monthlyCal(),'#2563eb','%');bar('patMonthlyChart',monthlyPat(),'#16a34a','%');bar('topIssuesChart',topOpenIssues(),'#dc2626','');bar('areaChart',areaComp(),'#8b5cf6','%')}
function renderTopOverdue(){let rows=state.calibration.filter(x=>x.status==='Overdue').sort((a,b)=>(a.daysLeft||0)-(b.daysLeft||0)).slice(0,8);$('topOverdueTable').innerHTML=rows.length?rows.map(r=>`<tr><td><b>${esc(r.gaugeId)}</b></td><td>${esc(r.machineNo)}</td><td><span class="badge danger-b">${Math.abs(r.daysLeft)} days</span></td></tr>`).join(''):'<tr><td>No overdue gauges</td></tr>'}
function monthlyCal(){let g={};state.calibration.forEach(r=>{let d=parseDate(r.dueDate);if(!d)return;let k=d.toLocaleString('en',{month:'short'});g[k]||={t:0,o:0};g[k].t++;if(r.status!=='Overdue')g[k].o++});return Object.entries(g).map(([label,v])=>({label,value:pct(v.o,v.t)}))}
function monthlyPat(){return MONTHS.map(m=>({label:m,value:calcPatByMonth(WEEKS.filter(w=>w.startsWith(m)))}))}
function topOpenIssues(){let map={};reportData().open.forEach(r=>{let key=r['Check Points'].slice(0,32);map[key]=(map[key]||0)+1});return Object.entries(map).sort((a,b)=>b[1]-a[1]).slice(0,8).map(([label,value])=>({label,value}))}
function areaComp(){let m={};state.calibration.forEach(r=>{m[r.area]||={t:0,o:0};m[r.area].t++;if(r.status!=='Overdue')m[r.area].o++});return Object.entries(m).map(([label,v])=>({label,value:pct(v.o,v.t)}))}
function bar(id,data,color,suffix){let c=$(id);if(!c)return;let ctx=c.getContext('2d'),dpr=devicePixelRatio||1,rect=c.getBoundingClientRect();c.width=rect.width*dpr;c.height=rect.height*dpr;ctx.scale(dpr,dpr);ctx.clearRect(0,0,rect.width,rect.height);if(!data.length){ctx.fillStyle='#64748b';ctx.fillText('No data available',20,40);return}let pad=40,max=Math.max(...data.map(d=>d.value),1),bw=Math.max(18,(rect.width-pad*2)/data.length-12);data.forEach((d,i)=>{let x=pad+i*((rect.width-pad*2)/data.length)+6,h=(rect.height-90)*(d.value/max),y=rect.height-50-h;ctx.fillStyle=color;ctx.beginPath();ctx.roundRect(x,y,bw,h,8);ctx.fill();ctx.fillStyle='#334155';ctx.font='11px Segoe UI';ctx.textAlign='center';ctx.fillText(String(d.label).slice(0,18),x+bw/2,rect.height-24);ctx.fillStyle='#0f172a';ctx.font='bold 12px Segoe UI';ctx.fillText(d.value+suffix,x+bw/2,Math.max(14,y-7))})}
function patLatestCompliance(){let ok=0,total=0;state.patrolling.forEach(r=>{let s=latestStatus(r);if(s.status!=='NOT_DONE'){total++;if(s.status==='OK')ok++}});return pct(ok,total)}
function calcPatByMonth(wks){let ok=0,total=0;state.patrolling.forEach(r=>{let s=latestStatus(r,wks);if(s.status!=='NOT_DONE'){total++;if(s.status==='OK')ok++}});return pct(ok,total)}
function weekCell(v){return`<td class="${v==='O'?'week-ok':v==='X'?'week-x':'week-blank'}">${v||'-'}</td>`}
function pager(p,total,page){let pages=Math.max(1,Math.ceil(total/state.pageSize));$(p+'PageInfo').textContent=`Page ${page} / ${pages}`;$(p+'Prev').disabled=page<=1;$(p+'Next').disabled=page>=pages}
function csvToObjects(text){let rows=[],row=[],field='',quote=false;for(let i=0;i<text.length;i++){let c=text[i],n=text[i+1];if(c==='"'&&quote&&n==='"'){field+='"';i++}else if(c==='"'){quote=!quote}else if(c===','&&!quote){row.push(field);field=''}else if((c==='\n'||c==='\r')&&!quote){if(c==='\r'&&n==='\n')i++;row.push(field);field='';if(row.some(x=>x!==''))rows.push(row);row=[]}else field+=c}if(field||row.length){row.push(field);rows.push(row)}let h=rows.shift().map(x=>x.trim().replace(/^\ufeff/,''));return rows.map(r=>Object.fromEntries(h.map((k,i)=>[k,(r[i]||'').trim()]))) }
function validate(rows,req,name){if(!rows.length)return{name,ok:false,message:name+' upload rejected. No data rows.'};let miss=req.filter(c=>!Object.keys(rows[0]).includes(c));return miss.length?{ok:false,message:name+' missing columns: '+miss.join(', ')}:{ok:true,message:name+' upload OK. Rows: '+rows.length}}
async function handleUpload(e,type){let file=e.target.files[0];if(!file)return;let text=await file.text(),rows=file.name.toLowerCase().endsWith('.json')?JSON.parse(text):csvToObjects(text);let res=type==='cal'?validate(rows,CAL_HEADERS,'Calibration'):validate(rows,['Machine','Sl.No','Check Points','Checking Method'],'Patrolling');log(res.message);if(!res.ok)return;if(type==='cal'){state.calibration=rows.map(normCal);$('calUploadStatus').textContent=rows.length+' loaded'}else{state.patrolling=rows.map(normPat);$('patUploadStatus').textContent=rows.length+' loaded'}state.filteredCal=state.calibration;state.filteredPat=state.patrolling;renderAll()}
async function syncCollection(name,rows,keyFn){if(!fb||!state.user){log('Login and Firebase config required.');return}let count=0;for(let i=0;i<rows.length;i+=450){let batch=writeBatch(fb.db);rows.slice(i,i+450).forEach(r=>batch.set(doc(fb.db,name,keyFn(r)),{...r,updatedAt:serverTimestamp(),updatedBy:state.user.email},{merge:true}));await batch.commit();count+=rows.slice(i,i+450).length;log(`Synced ${count}/${rows.length} to ${name}`)}await addDoc(collection(fb.db,'auditLogs'),{action:'bulkSync',collection:name,count:rows.length,user:state.user.email,createdAt:serverTimestamp()});}
async function uploadCertificates(){if(!fb||!state.user){log('Login and Firebase config required.');return}let files=[...$('certUpload').files];for(let f of files){let fileRef=ref(fb.storage,'certificates/'+f.name);await uploadBytes(fileRef,f);let url=await getDownloadURL(fileRef);log('Uploaded '+f.name+' => '+url)} }
function downloadText(n,t){let a=document.createElement('a');a.href=URL.createObjectURL(new Blob([t],{type:'text/plain'}));a.download=n;a.click();URL.revokeObjectURL(a.href)}
function applyCalFilters(){let g=$('gaugeSearch').value.toLowerCase(),m=$('machineSearchCal').value.toLowerCase(),a=$('areaSearch').value.toLowerCase(),c=$('certificateSearch').value.toLowerCase(),s=$('statusFilter').value;state.filteredCal=state.calibration.filter(r=>(!g||(`${r.gaugeId} ${r.gaugeDescription}`).toLowerCase().includes(g))&&(!m||String(r.machineNo).toLowerCase().includes(m))&&(!a||String(r.area).toLowerCase().includes(a))&&(!c||String(r.certificateNo).toLowerCase().includes(c))&&(!s||r.status===s));state.calPage=1;renderAll()}
function applyPatFilters(){let m=$('machineSearchPat').value.toLowerCase(),cp=$('checkpointSearch').value.toLowerCase(),st=$('patStatusFilter').value,mo=$('monthFilter').value,wk=mo?WEEKS.filter(w=>w.startsWith(mo)):WEEKS;state.filteredPat=state.patrolling.filter(r=>{let ls=latestStatus(r,wk);return(!m||r.Machine.toLowerCase().includes(m))&&(!cp||r['Check Points'].toLowerCase().includes(cp))&&(!st||ls.status===st)});state.patPage=1;renderAll()}
function setup(){document.querySelectorAll('.tab[data-tab]').forEach(b=>b.onclick=()=>{document.querySelectorAll('.tab').forEach(x=>x.classList.remove('active'));document.querySelectorAll('.section').forEach(x=>x.classList.remove('active'));b.classList.add('active');$(b.dataset.tab).classList.add('active');setTimeout(renderCharts,80)});['gaugeSearch','machineSearchCal','areaSearch','certificateSearch','statusFilter'].forEach(id=>$(id).addEventListener('input',debounce(applyCalFilters,200)));['machineSearchPat','checkpointSearch','patStatusFilter','monthFilter'].forEach(id=>$(id).addEventListener('input',debounce(applyPatFilters,200)));MONTHS.forEach(m=>$('monthFilter').insertAdjacentHTML('beforeend',`<option>${m}</option>`));$('calPrev').onclick=()=>{state.calPage--;renderCalTable()};$('calNext').onclick=()=>{state.calPage++;renderCalTable()};$('patPrev').onclick=()=>{state.patPage--;renderPatTable()};$('patNext').onclick=()=>{state.patPage++;renderPatTable()};$('toggleWeeks').onclick=()=>{showAllWeeks=!showAllWeeks;renderPatHeader();renderPatTable()};$('calUpload').onchange=e=>handleUpload(e,'cal');$('patUpload').onchange=e=>handleUpload(e,'pat');$('exportCalJson').onclick=()=>downloadText('calibration.json',JSON.stringify(state.calibration,null,2));$('exportPatJson').onclick=()=>downloadText('patrolling.json',JSON.stringify(state.patrolling,null,2));$('downloadCalCsv').onclick=()=>downloadText('calibration_template.csv',CAL_HEADERS.join(',')+'\n');$('downloadPatCsv').onclick=()=>downloadText('patrolling_template.csv',PAT_HEADERS.join(',')+'\n');$('syncCalFirestore').onclick=()=>syncCollection('calibration',state.calibration,r=>String(r.gaugeId||crypto.randomUUID()).replaceAll('/','_'));$('syncPatFirestore').onclick=()=>syncCollection('patrolling',state.patrolling,r=>String((r.Machine||'M')+'_'+(r['Sl.No']||crypto.randomUUID())).replaceAll('/','_'));$('uploadCerts').onclick=uploadCertificates;$('loginBtn').onclick=async()=>{if(!fb){log('Firebase config not filled.');return}await signInWithEmailAndPassword(fb.auth,$('email').value,$('password').value)};$('logoutBtn').onclick=async()=>{if(fb)await signOut(fb.auth)};if(fb){onAuthStateChanged(fb.auth,async u=>{state.user=u;$('userBadge').textContent=u?'Logged in: '+u.email:'Not logged in';await loadData()})}else{loadData()}}
window.addEventListener('resize',debounce(renderCharts,150));setup();
'''
(base/'assets/js/app.js').write_text(app_js, encoding='utf-8')

# Firebase rules and config
(base/'firebase/firestore.rules').write_text(r'''rules_version = '2';
service cloud.firestore {
  match /databases/{database}/documents {
    function signedIn() { return request.auth != null; }
    match /calibration/{docId} { allow read: if signedIn(); allow write: if signedIn(); }
    match /patrolling/{docId} { allow read: if signedIn(); allow write: if signedIn(); }
    match /auditLogs/{docId} { allow read, create: if signedIn(); allow update, delete: if false; }
    match /users/{uid} { allow read, write: if signedIn() && request.auth.uid == uid; }
  }
}
''', encoding='utf-8')
(base/'firebase/storage.rules').write_text(r'''rules_version = '2';
service firebase.storage {
  match /b/{bucket}/o {
    match /certificates/{fileName} {
      allow read: if request.auth != null;
      allow write: if request.auth != null && request.resource.size < 20 * 1024 * 1024;
    }
  }
}
''', encoding='utf-8')
(base/'firebase/firebase.json').write_text(json.dumps({"firestore":{"rules":"firestore.rules"},"storage":{"rules":"storage.rules"},"hosting":{"public":".","ignore":["firebase.json","**/.*","**/node_modules/**"]}}, indent=2), encoding='utf-8')

(base/'tools/excel_to_json.py').write_text('''import pandas as pd
from pathlib import Path
INPUT_FILE = "Compliance_Master.xlsx"
OUTPUT_DIR = Path("data")
OUTPUT_DIR.mkdir(exist_ok=True)
cal = pd.read_excel(INPUT_FILE, sheet_name="Calibration", engine="openpyxl").fillna("")
pat = pd.read_excel(INPUT_FILE, sheet_name="Patrolling", engine="openpyxl").fillna("")
cal.to_json(OUTPUT_DIR / "calibration.json", orient="records", indent=2, force_ascii=False)
pat.to_json(OUTPUT_DIR / "patrolling.json", orient="records", indent=2, force_ascii=False)
print("Generated JSON files in data folder")
''', encoding='utf-8')

readme = '''# Compliance Portal Firebase + GitHub Pages Bundle v5

This is the complete hosted portal bundle for Calibration Management and Patrolling Management.

## Included
- GitHub Pages frontend
- Firebase Auth login
- Firestore live database sync
- Firebase Storage certificate PDF upload
- Local JSON fallback if Firebase is not configured
- Calibration and patrolling Excel templates
- CSV templates
- Firestore and Storage security rule files

## Quick Hosting on GitHub Pages
1. Extract this zip.
2. Upload all files to GitHub repository root.
3. Open Settings > Pages.
4. Select Deploy from branch > main > root.
5. Open the GitHub Pages URL.

## Firebase Setup
1. Go to Firebase Console and create a project.
2. Add a Web App.
3. Copy Firebase web config.
4. Paste config into assets/js/firebase-config.js.
5. Enable Authentication > Email/Password.
6. Create your admin user in Firebase Authentication.
7. Create Firestore Database.
8. Create Storage.
9. Add GitHub Pages URL in Authentication > Settings > Authorized domains.
10. Upload firebase/firestore.rules and firebase/storage.rules rules manually in Firebase Console, or use Firebase CLI.

## Firestore Collections
- calibration
- patrolling
- auditLogs
- users

## Bulk Upload to Firestore
1. Login as admin in Bulk/Admin tab.
2. Upload calibration CSV/JSON.
3. Click Sync Calibration to Firestore.
4. Upload patrolling CSV/JSON.
5. Click Sync Patrolling to Firestore.
6. Upload certificate PDFs with Upload Certificates.

## Important
If Firebase is not configured or user is not logged in, portal automatically uses data/calibration.json and data/patrolling.json.
'''
(base/'README.md').write_text(readme, encoding='utf-8')
(base/'FIREBASE_SETUP_STEPS.md').write_text(readme, encoding='utf-8')
(base/'certificates/PUT_CERTIFICATE_PDFS_HERE.txt').write_text('If using GitHub-only certificate links, place PDFs here. With Firebase Storage, upload PDFs from Bulk/Admin tab.', encoding='utf-8')

zip_path=Path('/mnt/data/Compliance_Portal_Firebase_GitHub_Bundle_v5.zip')
if zip_path.exists(): zip_path.unlink()
with zipfile.ZipFile(zip_path,'w',zipfile.ZIP_DEFLATED) as z:
    for f in base.rglob('*'):
        z.write(f, f.relative_to(base))

with zipfile.ZipFile(zip_path) as z:
    names=z.namelist()
    required=['index.html','assets/js/app.js','assets/js/firebase-config.js','assets/css/style.css','firebase/firestore.rules','firebase/storage.rules','data/calibration.json','data/patrolling.json','README.md']
    missing=[r for r in required if r not in names]
print(json.dumps({'zip':str(zip_path),'size_kb':round(zip_path.stat().st_size/1024,1),'files':len(names),'missing':missing}, indent=2))
